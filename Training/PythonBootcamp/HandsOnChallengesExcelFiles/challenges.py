# Challenge #1
#
# Create a Python script that uses openpyxl and generates an Excel file with a single sheet and the below content:

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active


rows = (
    ('Year', 'Sales'),
    (2017, 150000),
    (2018, 180000),
    (2019, 210000),
    (2020, 125000),

)

for r in rows:
    sheet.append(r)


wb.save('sales.xlsx')

# Challenge #2
#
# Change the solution from the previous challenge and set the sheet title to COMPANY SALES.

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'COMPANY SALES'

rows = (
    ('Year', 'Sales'),
    (2017, 150000),
    (2018, 180000),
    (2019, 210000),
    (2020, 125000),

)

for r in rows:
    sheet.append(r)


wb.save('sales.xlsx')

# Challenge #3
#
# Consider this Excel file.
#
# a) Add a new sheet called VAT.
#
# b) Read the Excel file in Python, calculate the VAT if the VAT rate is 15%, and write the data into the new sheet.
#
# c) Save the Excel file as a new file called sales_and_vat.xlsx
import openpyxl
wb = openpyxl.load_workbook('sales.xlsx')

sheet = wb.active

# read the entire content in a list
items = list()
for row in sheet.values:
    items.append(row)

# print(items)

# calculating the vat values
vat = list()
for row in items[1:]:
    element = (row[0], row[1] * 0.15)
    vat.append(element)

# print(vat)

# add a new sheet called vat
wb.create_sheet('VAT')
sheet = wb['VAT']

# add a header (column names)
sheet['A1'] = 'Year'
sheet['B1'] = 'VAT'

# appending to the new sheet
for row in vat:
    sheet.append(row)

wb.save('sales_and_vat.xlsx')

# Challenge #4
#
# Consider this Excel file.
#
# Add an Excel formula into the B6 cell so that it calculates the total sales.
#
# Add the string Total Sales into cell C6.
#
import openpyxl
# sales2.xlsx -> https://drive.google.com/open?id=1Yw0j7v49-Mcrve6vN3h6yLTMxkSmdqqx
wb = openpyxl.load_workbook('sales2.xlsx')

sheet = wb.active


cell = sheet['B6']
cell.value = '=sum(B2:B5)'


cell = sheet['C6']

cell.value = 'Total Sales'

wb.save('sales3.xlsx')

# Consider the Excel file generated in the previous challenge.
#
# Format the cells B6 and C6 in the following way: Tahoma font family, size 16, color red, and bold.

import openpyxl
from openpyxl.styles import *

# sales3.xlsx -> https://drive.google.com/open?id=1a5iJfQORxz8OROT15m-bzeY2gNd26OlA
wb = openpyxl.load_workbook('sales3.xlsx')

sheet = wb.active


font = Font(name='Tahoma', size=16, color=colors.RED, bold=True, italic=False, strike=False)
cell_b6 = sheet['B6']
cell_b6.font = font

cell_c6 = sheet['c6']
cell_c6.font = font


wb.save('sales3.xlsx')

#  Challenge 7

def csv2excel(filein, fileout, delim=','):
    import openpyxl, csv
    with open(filein, 'r') as f:
        reader = csv.reader(f, delimiter=delim)

        wb = openpyxl.Workbook()
        sheet = wb.active
        for row in reader:
            sheet.append(row)

        wb.save(fileout)


# Challenge #8
#
# Create a Python script that contains a function called excel2csv() which exports any Excel file to CSV one.
#
# The function will be called like this: excel2csv('books.xlsx', 'booklist.csv')
#
# It will read the Excel file called books.xlsx and export it to booklist.csv
# csv2excel('people3.csv', 'teachers.xlsx')

def excel2csv(filein, fileout):
    import openpyxl, csv
    wb = openpyxl.load_workbook(filein)
    sheet = wb.active
    with open(fileout, 'w') as f:
        writer = csv.writer(f)
        for row in sheet.values:
            writer.writerow(row)

excel2csv('books.xlsx', 'booklist.csv')